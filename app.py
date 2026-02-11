# -*- coding: utf-8 -*-
import io
import time
import shutil
import tempfile
from datetime import datetime

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from pdf_contracts import (
    HEADERS,
    parse_contract,
    extract_raw_and_normalized_text,
    calc_quality,
)

APP_TITLE = "📄 تحويل عقود PDF إلى Excel (قوي + تقرير)"
OUTPUT_FILE_NAME = "Employees_Data.xlsx"
SHEET_MAIN = "الموظفين"
SHEET_LOGS = "Logs"

st.set_page_config(page_title="PDF → Excel (عقود الموظفين)", page_icon="📄", layout="wide")
st.title(APP_TITLE)

st.write(
    "ارفع ملفات PDF (أي عدد). يتم استخراج البيانات ووضع كل موظف في سطر واحد داخل Excel.\n"
    "أي حقل غير موجود في العقد سيبقى فارغ. وإذا ملف واحد فيه مشكلة، العملية تكمل للباقي."
)

with st.expander("⚙️ إعدادات", expanded=False):
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        include_logs_sheet = st.checkbox("إضافة ورقة Logs في Excel", value=True)
    with c2:
        enable_debug = st.checkbox("Debug: عرض النص الخام + بعد التطبيع", value=True)
    with c3:
        show_quality_table = st.checkbox("عرض جدول جودة الاستخراج", value=True)
    with c4:
        cleanup_delay = st.slider("ثواني قبل تنظيف الملفات المؤقتة", 0, 8, 2)

uploaded = st.file_uploader("ارفع ملفات PDF هنا", type=["pdf"], accept_multiple_files=True)

def _auto_width(ws, max_width=70, min_width=10):
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value or ""
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(min_width, max_len + 2), max_width)

def build_excel_bytes(rows, logs=None, include_logs=True):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_MAIN

    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(h, "") if row.get(h, "") is not None else "" for h in HEADERS])

    ws.freeze_panes = "A2"
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEADERS)):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _auto_width(ws)

    if include_logs:
        ws2 = wb.create_sheet(SHEET_LOGS)
        ws2.append([
            "timestamp", "file_name", "status",
            "filled_fields", "total_fields", "quality_%", "missing_fields",
            "note"
        ])
        for c in range(1, 9):
            cell = ws2.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if logs:
            for item in logs:
                ws2.append([
                    item.get("timestamp",""),
                    item.get("file_name",""),
                    item.get("status",""),
                    item.get("filled_fields",""),
                    item.get("total_fields",""),
                    item.get("quality_pct",""),
                    item.get("missing_fields",""),
                    item.get("note",""),
                ])

        ws2.freeze_panes = "A2"
        for row_cells in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=8):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        _auto_width(ws2, max_width=90)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def safe_lines(s: str, n=120) -> str:
    if not s:
        return ""
    return "\n".join(s.splitlines()[:n])

def process_files(files):
    rows = []
    logs = []
    debug_items = []
    report_lines = []

    total_files = len(files)
    progress = st.progress(0)
    status = st.empty()

    for i, f in enumerate(files, start=1):
        status.write(f"جارٍ معالجة الملف {i}/{total_files}: **{f.name}**")
        ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

        try:
            pdf_bytes = f.read()
            if not pdf_bytes or len(pdf_bytes) < 50:
                row = {h: "" for h in HEADERS}
                filled, total, pct, missing = calc_quality(row)

                rows.append(row)
                logs.append({
                    "timestamp": ts,
                    "file_name": f.name,
                    "status": "SKIPPED",
                    "filled_fields": filled,
                    "total_fields": total,
                    "quality_pct": pct,
                    "missing_fields": ", ".join(missing[:10]) + (" ..." if len(missing) > 10 else ""),
                    "note": "File empty/too small"
                })

                report_lines.append(f"- {f.name}: SKIPPED (empty)")
                debug_items.append({"file": f.name, "raw": "", "norm": "", "note": "empty"})
            else:
                raw_text, norm_text = extract_raw_and_normalized_text(pdf_bytes)
                data = parse_contract(norm_text) or {}
                row = {h: (data.get(h, "") if data.get(h, "") is not None else "") for h in HEADERS}

                filled, total, pct, missing = calc_quality(row)

                status_label = "OK" if pct >= 35 else "LOW_QUALITY"
                note = "Parsed successfully" if status_label == "OK" else "Low filled fields; check Debug text"

                rows.append(row)
                logs.append({
                    "timestamp": ts,
                    "file_name": f.name,
                    "status": status_label,
                    "filled_fields": filled,
                    "total_fields": total,
                    "quality_pct": pct,
                    "missing_fields": ", ".join(missing[:10]) + (" ..." if len(missing) > 10 else ""),
                    "note": note
                })

                report_lines.append(f"- {f.name}: {status_label} | Quality {pct}% | Missing {len(missing)} fields")

                if enable_debug:
                    debug_items.append({
                        "file": f.name,
                        "raw": safe_lines(raw_text, 80),
                        "norm": safe_lines(norm_text, 120),
                        "note": f"Quality {pct}%"
                    })

        except Exception as e:
            row = {h: "" for h in HEADERS}
            filled, total, pct, missing = calc_quality(row)

            rows.append(row)
            logs.append({
                "timestamp": ts,
                "file_name": f.name,
                "status": "ERROR",
                "filled_fields": filled,
                "total_fields": total,
                "quality_pct": pct,
                "missing_fields": ", ".join(missing[:10]) + (" ..." if len(missing) > 10 else ""),
                "note": f"{type(e).__name__}: {str(e)}"
            })
            report_lines.append(f"- {f.name}: ERROR -> {type(e).__name__}: {str(e)}")
            debug_items.append({"file": f.name, "raw": "", "norm": "", "note": f"ERROR: {e}"})

        progress.progress(int(i / total_files * 100))

    status.write("✅ انتهت المعالجة.")
    report_text = "PDF Contracts Extraction Report\n" + "\n".join(report_lines)
    return rows, logs, debug_items, report_text

if uploaded:
    st.info(f"عدد الملفات المرفوعة: **{len(uploaded)}**")
    run = st.button("⚙️ تحويل إلى Excel", type="primary")
else:
    run = False

if run:
    temp_dir = tempfile.mkdtemp(prefix="pdf_to_excel_")
    try:
        rows, logs, debug_items, report_text = process_files(uploaded)

        if show_quality_table:
            st.subheader("📊 جودة الاستخراج لكل ملف")
            for item in logs:
                st.write(
                    f"- **{item['file_name']}** | Status: `{item['status']}` | "
                    f"Filled: {item['filled_fields']}/{item['total_fields']} | "
                    f"Quality: **{item['quality_pct']}%** | Missing: {item['missing_fields']}"
                )

        excel_bytes = build_excel_bytes(rows, logs=logs, include_logs=include_logs_sheet)

        st.success("✅ تم تجهيز ملف Excel بنجاح!")
        st.download_button(
            label="⬇️ تنزيل Employees_Data.xlsx",
            data=excel_bytes,
            file_name=OUTPUT_FILE_NAME,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.download_button(
            label="⬇️ تنزيل تقرير مفصل (TXT)",
            data=report_text.encode("utf-8"),
            file_name="Extraction_Report.txt",
            mime="text/plain",
        )

        if enable_debug:
            st.subheader("🧪 Debug لكل ملف (خام + بعد التطبيع)")
            st.caption("لو حقل ما يطلع، افتح الملف وشوف النص بعد التطبيع — هذا هو اللي نعتمد عليه في الاستخراج.")
            for d in debug_items:
                with st.expander(f"📄 {d['file']} — {d['note']}", expanded=False):
                    st.text_area("RAW (first 80 lines)", d.get("raw",""), height=210)
                    st.text_area("NORMALIZED (first 120 lines)", d.get("norm",""), height=260)

        st.info("🧹 سيتم حذف الملفات المؤقتة تلقائيًا.")
        time.sleep(int(cleanup_delay))

    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)
